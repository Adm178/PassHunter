from tkinter import *
from scan import find_passwords
from tkinter import filedialog
from tkinter import messagebox
import os
import openpyxl

# Выбор дикректории сканирования
directory = "C:/"
def get_directory(directory_entry):
    global directory
    directory = filedialog.askdirectory()
    directory_entry.delete(0, END)  # очищаем строку ввода
    directory_entry.insert(0, directory)  # вставляем адрес выбранной директории

directory2 = " "
def get_directory2(directory_entry2):
    global directory2
    directory2 = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
    directory_entry2.delete(0, END)  # очищаем строку ввода
    directory_entry2.insert(0, directory2)  # вставляем адрес выбранной директории

# Запуск сканирования
def scan(tree, option):
    if directory == "C:/":
        messagebox.showerror("Ошибка", "Директория не выбрана")
    else:
        tree.delete(*tree.get_children())
        result = find_passwords(directory, directory2, option)
        if result != " ":
            for i, (password, path, entropy, passwordmeter, probability) in enumerate(result, 1):
                tree.insert("", "end", text=str(i), values=(i, password, path, entropy, passwordmeter, probability))

# Сортировка столбцов по возрастанию и убыванию
reverse=False
def on_header_click(event, tree):
    column = tree.identify_column(event.x)  # Определение столбца, по которому было произведено нажатие
    global reverse
    reverse = not reverse
    data = [(tree.set(child, column), child) for child in tree.get_children('')]
    data.sort(key=lambda x: num_key(x[0]), reverse=reverse)  # Сортировка данных по указанному столбцу
    for index, (_, child) in enumerate(data):
        tree.move(child, '', index)  # Перемещение элементов в таблице в новый порядок

# Перевод строки в число или 0 при ошибке
def num_key(num_str):
    try:
        return float(num_str)
    except ValueError:
        return 0

# Экспорт данных в файл
def export_to_txt(tree):
    filename = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=
            [("Text Files", "*.txt"),("Excel Files", "*.xlsx")])
    if not filename:
        return  # Если пользователь не указал имя файла, прерываем операцию
    else:
        _, extension = os.path.splitext(filename)
        if extension == ".txt":
            with open(filename, "w") as file:
                for child in tree.get_children():
                    values = tree.item(child)["values"]
                    line = "\t".join(str(value) for value in values)  # Соединяем значения через табуляцию
                    file.write(line + "\n")  # Записываем строку в файл
        else:
            # Создание нового Excel-файла
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Получение заголовков столбцов из Treeview
            columns = tree['columns']
            for i, column in enumerate(columns):
                column_heading = tree.heading(column)['text']
                sheet.cell(row=1, column=i + 1, value=column_heading)
            # Запись данных из Treeview в Excel
            for row_id in tree.get_children():
                row = tree.item(row_id)['values']
                for i, value in enumerate(row):
                    sheet.cell(row=str(row_id+2), column=str(i + 1), value=value)
            # Сохранение файла Excel
            workbook.save(filename)

    messagebox.showinfo("Экспорт завершен", "Данные успешно сохранены.")

# Копирование значения выделенной ячейки
def copy_to_clipboard(tree, root):
    try:
        row_id = tree.selection()[0]  # получаем ID выделенной строки
        col_index = tree.identify_column(coord_x)  # Определение столбца, по которому было произведено нажатие
        col_id = int(col_index.replace('#', ''))
        value = tree.item(row_id)['values'][col_id-1]  # получаем значение ячейки
        root.clipboard_clear()  # очищаем буфер обмена
        root.clipboard_append(value)  # добавляем значение в буфер обмена
    except IndexError:
        print("Ячейка не выбрана")

# Создание контекстного меню
def f_menu(event, menu):
    global coord_x
    coord_x = event.x
    menu.post(event.x_root, event.y_root)

def state(selected_option, option2_button):
    if selected_option.get() != "option2":
        option2_button.config(state="disabled")
    else:
        option2_button.config(state="normal")