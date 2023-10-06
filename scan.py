import os
import math
from collections import Counter
import set_wind
from docx import Document
from openpyxl import load_workbook
from tkinter import messagebox
import passwordmeter
import tensorflow as tf
import numpy as np
import re
model = tf.keras.models.load_model('./model-lstm')

# извлечено из токенизатора fit Keras, поэтому нам не нужен Keras / Tensorflow в качестве требования
CHAR_DICT = {'<UNK>': 1, 'e': 2, 'i': 3, 'a': 4, 'n': 5, 't': 6, 'r': 7, 'o': 8, 's': 9, 'c': 10, 'l': 11, 'A': 12,
             'E': 13, 'd': 14, 'u': 15, 'm': 16, 'p': 17, 'I': 18, 'S': 19, 'R': 20, 'O': 21, 'N': 22, 'g': 23, 'T': 24,
             '-': 25, 'L': 26, 'h': 27, 'y': 28, 'C': 29, 'b': 30, 'f': 31, 'M': 32, 'v': 33, 'D': 34, '1': 35, 'U': 36,
             'H': 37, 'P': 38, 'k': 39, '2': 40, '0': 41, 'B': 42, 'G': 43, 'w': 44, 'Y': 45, 'K': 46, '3': 47, '9': 48,
             'F': 49, '.': 50, ',': 51, '4': 52, '8': 53, 'V': 54, '5': 55, '7': 56, '6': 57, 'W': 58, 'j': 59, 'x': 60,
             'z': 61, 'J': 62, 'q': 63, 'Z': 64, '_': 65, "'": 66, ':': 67, 'X': 68, 'Q': 69, '/': 70, ')': 71, '(': 72,
             '"': 73, '!': 74, ';': 75, '*': 76, '@': 77, '\\': 78, ']': 79, '?': 80, '[': 81, '<': 82, '>': 83,
             '=': 84, '#': 85, '&': 86, '$': 87, '+': 88, '%': 89, '`': 90, '~': 91, '^': 92, '{': 93, '}': 94, '|': 95}

# регулярное выражение для слов, в которых присутствует хотя бы одна цифра,
# хотя бы одна латинская буква (в нижнем или верхнем регистре)
# хотя бы один специальный символ и что длина строки от 7 до 32 символов
PASSWORD_REGEX = re.compile('^(?=.*[0-9])(?=.*[a-z])(?=.*[A-Z])(?=.*[~!@#$%^&*_\-+=`|\()\{\}[\]:;"\'<>,.?\/])(?=.*\d).{7,32}$')

# Поиск всех файлов в директории и во всех подпапках
def get_files_recursive(directory):
    file_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_path = file_path.replace("\\", "/")  # Замена обратных слэшей на прямые слэши
            file_list.append(file_path)
    return file_list
def find_passwords(directory, directory2, option):
    global result

    files = get_files_recursive(directory)
    result = []
    if option.get() == "option2":
        if directory2 == " ":
            messagebox.showerror("Ошибка", "Словарь не выбран")
            return " "
        else:
            with open(directory2, 'r') as f:
                wordbook = f.read()
    else:
        wordbook = " "

    # Проход по каждому файлу и проверка его содержимого на наличие пароля
    for file in files:
        content = ""
        if file.endswith(('.txt')):
            if set_wind.check1:
                with open(os.path.join(directory, file), 'r') as f:
                    content = f.read()

        elif file.endswith(('.docx', '.doc')):
            if set_wind.check2:
                doc = Document(os.path.join(directory, file))
                paragraphs = [p.text for p in doc.paragraphs]
                content = "\n".join(paragraphs)

        elif file.endswith(('.xlsx', '.xls')):
            if set_wind.check3:
                wb = load_workbook(os.path.join(directory, file))
                sheet = wb.active
                rows = sheet.iter_rows(values_only=True)
                for row in rows:
                    content += " ".join(str(cell) for cell in row) + "\n"

        elif file.endswith(('.config', '.ini', '.xml', '.yaml', '.log')):
            if set_wind.check4:
                with open(os.path.join(directory, file), 'r') as f:
                    content = f.read()

        words = content.split()  # Разделение текста на слова

        if option.get() == "option1":
            find_entropy(words, file)
        elif option.get() == "option2":
            find_wordbook(words, file, wordbook)
        elif option.get() == "option3":
            find_AI(words, file)
        # print(len(result))
    return result

# Вычисление энтропии текста
def has_password(text):
    freq = Counter(text)
    probs = [count / len(text) for count in freq.values()]
    entropy = - sum(p * math.log(p, 2) for p in probs)
    return entropy
def find_entropy(words, file):
    # Поиск по регулярному выражению
    if set_wind.check7 == True:
        for x, word in enumerate(words):
            if PASSWORD_REGEX.match(word):
                inner_list = [words[x], file, None, None, None]
                result.append(inner_list)

    if set_wind.check5 == True:
        # Поиск по ключевым словам с частичным совпадением
        if set_wind.check6 == True:
            for word in set_wind.array:
                for i in range(len(words)):
                    if word in words[i]:
                        inner_list = [words[i + 1] if i + 1 < len(words) else None, file, None, None, None]
                        result.append(inner_list)
        # Поиск по ключевым словам с полным совпадением
        else:
            for regex in set_wind.array:
                for i in range(len(words)):
                    if regex == words[i]:
                        inner_list = [words[i + 1] if i + 1 < len(words) else None, file, None, None, None]
                        result.append(inner_list)
    for word in words:
        if float(set_wind.input_text3) <= len(word) <= float(set_wind.input_text4):
            passwordmeters, _ = passwordmeter.test(word)
            if passwordmeters*10 >= float(set_wind.input_text5):
                entropy = round(has_password(word), 3)
                # Проверка наличия пароля по заданному порогу энтропии
                if float(set_wind.input_text1) < entropy < float(set_wind.input_text2):
                    tokenized_word = tokenize(word)
                    word_prob = np.array([tokenized_word])
                    pred = model.predict(word_prob)
                    inner_list = [word, file, entropy, round(passwordmeters*10, 3), round(float(pred[0]), 3)]
                    result.append(inner_list)

def find_wordbook(words, file, wordbook):
    if set_wind.check7 == True:
        for x, word in enumerate(words):
            if PASSWORD_REGEX.match(word):
                inner_list = [words[x], file, None, None, None]
                result.append(inner_list)

    if set_wind.check5 == True:
        # Ищем по ключевым словам с частичным совпадением
        if set_wind.check6 == True:
            for word in set_wind.array:
                for i in range(len(words)):
                    if word in words[i]:
                        inner_list = [words[i + 1] if i + 1 < len(words) else None, file, None, None, None]
                        result.append(inner_list)
        # Ищем по ключевым словам с полным совпадением
        else:
            for word in set_wind.array:
                for i in range(len(words)):
                    if word == words[i]:
                        inner_list = [words[i + 1] if i + 1 < len(words) else None, file, None, None, None]
                        result.append(inner_list)

    words2 = wordbook.split()
    for word in words:
        for word2 in words2:
            entropy = round(has_password(word), 3)
            if word == word2:
                passwordmeters, _ = passwordmeter.test(word)
                tokenized_word = tokenize(word)
                word_prob = np.array([tokenized_word])
                pred = model.predict(word_prob)
                inner_list = [word, file, entropy, round(passwordmeters*10, 3), round(float(pred[0]), 3)]
                result.append(inner_list)

# Поиск через нейросеть
def find_AI(text, file):
    # Преобразуем вводимые слова в последовательности и увеличиваем их до 32
    tokenized_words = [tokenize(word) for word in text]
    words = np.array(tokenized_words)
    # Размещаем маркированные слова в обслуживаемой модели
    pred = model.predict(words)

    # Определение вероятности, что слово является паролем
    pred_bool = (pred > set_wind.input_text6)
    # Используйте предсказания в качестве индексов для возвращаемых слов
    positive_indicies = np.where(pred_bool == True)
    for x in positive_indicies[0]:
        password = text[x]
        entropy = round(has_password(password), 3)
        passwordmeters, _ = passwordmeter.test(password)
        inner_list = [password, file, entropy, round(passwordmeters * 10, 3), round(float(pred[x]), 3)]
        result.append(inner_list)


def tokenize(word):  # Выделяет и дополняет введенное слово до нужной длины.
    global CHAR_DICT

    if len(word) < 7 or len(word) > 32:
        return [0] * 32
    else:
        seq = [CHAR_DICT[char] if char in CHAR_DICT else 1 for char in word]
        seq.extend([0] * (32 - len(seq)))
        return seq